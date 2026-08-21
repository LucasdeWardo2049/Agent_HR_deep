"""Google Drive/Sheets integration behind one Composio client."""

import asyncio
import base64
import hashlib
import mimetypes
import tempfile
import threading
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast
from urllib.parse import quote

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas import CandidateAssessment, CandidateProfile, JobProfile, ResumeFile
from app.settings import Settings, get_settings

PDF_MIME = "application/pdf"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
GOOGLE_DOC_MIME = "application/vnd.google-apps.document"
SUPPORTED_MIME_TYPES = (PDF_MIME, DOCX_MIME, GOOGLE_DOC_MIME)

CRITERION_TYPE_LABELS = {
    "education": "Formação",
    "language": "Idioma",
    "experience": "Experiência",
    "skill": "Competência",
    "certification": "Certificação",
    "other": "Outro",
}
CRITERION_STATUS_LABELS = {
    "supported": "Atendido",
    "partial": "Parcialmente atendido",
    "not_found": "Não encontrado",
    "unclear": "Não está claro",
}


class GoogleWorkspaceError(RuntimeError):
    pass


@dataclass(frozen=True)
class ReportArtifacts:
    google_sheet_url: str
    excel_url: str
    excel_file_id: str


@dataclass(frozen=True)
class ReportTables:
    summary: list[list[Any]]
    criteria: list[list[Any]]
    candidates: list[list[Any]]
    candidate_cv_urls: list[str]


def build_report_tables(
    job_profile: JobProfile,
    assessments: list[CandidateAssessment],
    profiles: list[CandidateProfile],
    public_app_url: str = "http://localhost:8000",
) -> ReportTables:
    profile_by_id = {profile.candidate_id: profile for profile in profiles}
    required_count = sum(criterion.required for criterion in job_profile.criteria)
    summary: list[list[Any]] = [
        ["Busca de Talentos", ""],
        ["Vaga", job_profile.title],
        ["Data", datetime.now(UTC).isoformat()],
        ["Candidatos analisados", len(assessments)],
        ["Critérios obrigatórios", required_count],
        ["Critérios desejáveis", len(job_profile.criteria) - required_count],
        [
            "Observação",
            "Resultados representam evidências dos currículos e exigem revisão humana.",
        ],
    ]
    criteria: list[list[Any]] = [["ID do critério", "Descrição", "Obrigatório", "Tipo de critério"]]
    criteria.extend(
        [
            criterion.id,
            criterion.description,
            "Sim" if criterion.required else "Não",
            CRITERION_TYPE_LABELS[criterion.criterion_type],
        ]
        for criterion in job_profile.criteria
    )
    candidates: list[list[Any]] = [
        [
            "Candidato",
            "Currículo",
            "Formação",
            "Idiomas",
            "Experiência relevante (anos)",
            "Competências",
            "Critérios obrigatórios atendidos",
            "Total de critérios obrigatórios",
            "Cobertura dos critérios",
            "Evidências",
            "Pontos a confirmar",
            "Resumo profissional",
        ]
    ]
    candidate_cv_urls: list[str] = []
    for assessment in sorted(assessments, key=lambda item: (item.candidate_name or "").casefold()):
        profile = profile_by_id[assessment.candidate_id]
        cv_url = f"{public_app_url.rstrip('/')}/api/v1/talent/candidates/{quote(profile.candidate_id, safe='')}/cv"
        candidate_cv_urls.append(cv_url)
        education = "; ".join(profile.education)
        languages = "; ".join(profile.languages)
        evidence = " | ".join(
            f"{criterion.criterion_id}: {CRITERION_STATUS_LABELS[criterion.status]}"
            + (f" ({'; '.join(criterion.evidence)})" if criterion.evidence else "")
            for criterion in assessment.criteria
        )
        candidates.append(
            [
                assessment.candidate_name or "Nome não informado",
                cv_url,
                education,
                languages,
                profile.relevant_experience_years,
                ", ".join(profile.skills),
                assessment.required_supported,
                assessment.required_total,
                assessment.criteria_coverage,
                evidence,
                "; ".join(assessment.points_to_confirm),
                assessment.professional_summary or "",
            ]
        )
    return ReportTables(
        summary=_safe_rows(summary),
        criteria=_safe_rows(criteria),
        candidates=_safe_rows(candidates),
        candidate_cv_urls=candidate_cv_urls,
    )


def create_xlsx(path: Path, tables: ReportTables) -> None:
    header_fill = PatternFill("solid", fgColor="E7194B")
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Inter", size=10, color="11132C")
    widths = {
        "Summary": [28, 72],
        "Criteria": [22, 64, 16, 22],
        "Candidates": [26, 14, 44, 28, 20, 44, 24, 22, 20, 72, 56, 64],
    }
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)
    for title, rows in (
        ("Summary", tables.summary),
        ("Criteria", tables.criteria),
        ("Candidates", tables.candidates),
    ):
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_number, width in enumerate(widths[title], start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=column_number).column_letter].width = width
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = "E7194B"
        sheet.freeze_panes = "A2"
        if title != "Summary":
            sheet.auto_filter.ref = sheet.dimensions
        if title == "Candidates":
            for row_number, cv_url in enumerate(tables.candidate_cv_urls, start=2):
                cell = sheet.cell(row=row_number, column=2)
                if cv_url:
                    cell.value = "Abrir currículo"
                    cell.hyperlink = cv_url
                    cell.style = "Hyperlink"
                    cell.font = Font(name="Inter", size=10, bold=True, color="E7194B", underline="single")
                sheet.cell(row=row_number, column=9).number_format = "0%"
    workbook.save(path)


class GoogleWorkspaceClient:
    """Execute Google tools through one user-scoped Composio Session."""

    def __init__(self, settings: Settings | None = None) -> None:
        self.settings = settings or get_settings()
        self.client: Any | None = None
        self.session: Any | None = None
        self.last_log_id: str | None = None
        self.last_request_id: str | None = None
        self._client_lock = threading.RLock()

    def _get_client(self) -> Any:
        if self.client is not None:
            return self.client
        with self._client_lock:
            if self.client is not None:
                return self.client
            self.settings.require_google()
            from composio import Composio

            temp_root = tempfile.gettempdir()
            self.client = Composio(
                api_key=cast(str, self.settings.composio_api_key),
                toolkit_versions={
                    "googledrive": self.settings.composio_googledrive_version,
                    "googlesheets": self.settings.composio_googlesheets_version,
                },
                timeout=self.settings.composio_request_timeout_seconds,
                max_retries=self.settings.composio_max_retries,
                dangerously_allow_auto_upload_download_files=True,
                file_download_dir=temp_root,
                file_upload_dirs=[temp_root],
            )
        return self.client

    def _get_session(self) -> Any:
        if self.session is not None:
            return self.session
        with self._client_lock:
            if self.session is not None:
                return self.session
            self.session = self._get_client().create(
                user_id=cast(str, self.settings.composio_user_id),
                toolkits=["googledrive", "googlesheets"],
                manage_connections=False,
                sandbox={"enable": False},
            )
        return self.session

    def _execute(self, slug: str, arguments: dict[str, Any], *, toolkit: str) -> Any:
        if toolkit not in {"googledrive", "googlesheets"}:
            raise ValueError(f"Unsupported Google toolkit: {toolkit}")
        self.last_request_id = None
        try:
            response = self._get_session().execute(
                slug,
                arguments=arguments,
            )
        except Exception as exc:
            error_body = getattr(exc, "body", None)
            request_id = _find_value(error_body, ("request_id", "requestId"))
            error_slug = _find_value(error_body, ("slug",))
            self.last_request_id = request_id
            request_suffix = f" (Composio request: {request_id})" if request_id else ""
            if error_slug == "ToolRouterV2_NoActiveConnection":
                message = (
                    "Nenhuma conexão Google ativa para o COMPOSIO_USER_ID configurado. "
                    f"Autorize novamente o toolkit {toolkit}{request_suffix}."
                )
            else:
                message = f"Falha ao executar {slug} no Composio{request_suffix}."
            raise GoogleWorkspaceError(message) from exc
        if isinstance(response, dict):
            raw = response
        elif hasattr(response, "model_dump"):
            raw = response.model_dump()
        else:
            raw = {
                "data": getattr(response, "data", None),
                "successful": getattr(response, "successful", True),
                "error": getattr(response, "error", None),
            }
        log_id = raw.get("log_id") or raw.get("logId")
        self.last_log_id = str(log_id) if log_id else None
        if raw.get("successful") is False or raw.get("error"):
            log_suffix = f" (Composio log: {self.last_log_id})" if self.last_log_id else ""
            raise GoogleWorkspaceError(f"{slug} failed{log_suffix}: {raw.get('error')}")
        return raw.get("data", raw)

    def _stage_file(self, path: Path, *, slug: str, toolkit: str) -> dict[str, str]:
        """Upload a temporary local file and return Composio's tool descriptor."""
        resolved = path.resolve(strict=True)
        temp_root = Path(tempfile.gettempdir()).resolve(strict=True)
        if not resolved.is_relative_to(temp_root):
            raise GoogleWorkspaceError("Only files inside the temporary directory may be uploaded")
        if not resolved.is_file():
            raise GoogleWorkspaceError(f"File is not readable: {resolved.name}")

        digest = hashlib.md5(usedforsecurity=False)
        with resolved.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        mimetype = mimetypes.guess_type(resolved.name)[0] or "application/octet-stream"
        upload = self._get_client().client.files.create_presigned_url(
            filename=resolved.name,
            md5=digest.hexdigest(),
            mimetype=mimetype,
            tool_slug=slug,
            toolkit_slug=toolkit,
        )
        try:
            with resolved.open("rb") as source:
                response = httpx.put(upload.new_presigned_url, content=source, timeout=60)
                response.raise_for_status()
        except (OSError, httpx.HTTPError) as exc:
            raise GoogleWorkspaceError(f"Could not stage {resolved.name} for upload") from exc
        return {"name": resolved.name, "mimetype": mimetype, "s3key": upload.key}

    async def list_resume_files(self) -> list[ResumeFile]:
        folder_id = self.settings.google_drive_talent_pool_folder_id
        mime_query = " or ".join(f"mimeType = '{mime}'" for mime in SUPPORTED_MIME_TYPES)
        query = f"'{folder_id}' in parents and trashed = false and ({mime_query})"
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_FIND_FILE",
            {"q": query},
            toolkit="googledrive",
        )
        records = _find_file_records(data)
        files: list[ResumeFile] = []
        for record in records:
            file_id = record.get("id") or record.get("file_id")
            name = record.get("name") or record.get("file_name")
            mime_type = record.get("mimeType") or record.get("mime_type")
            if not file_id or not name or mime_type not in SUPPORTED_MIME_TYPES:
                continue
            files.append(
                ResumeFile(
                    drive_file_id=str(file_id),
                    file_name=str(name),
                    mime_type=str(mime_type),
                    drive_url=record.get("webViewLink")
                    or record.get("web_view_link")
                    or f"https://drive.google.com/open?id={file_id}",
                    modified_time=record.get("modifiedTime") or record.get("modified_time"),
                )
            )
        return files

    async def download_resume(self, file: ResumeFile) -> bytes:
        arguments: dict[str, Any] = {"file_id": file.drive_file_id}
        if file.mime_type == GOOGLE_DOC_MIME:
            arguments["mime_type"] = "text/plain"
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_DOWNLOAD_FILE",
            arguments,
            toolkit="googledrive",
        )
        return await _read_download(data)

    async def upload_file(self, path: Path, folder_id: str) -> str:
        uploadable = await asyncio.to_thread(
            self._stage_file,
            path,
            slug="GOOGLEDRIVE_UPLOAD_FILE",
            toolkit="googledrive",
        )
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_UPLOAD_FILE",
            {"file_to_upload": uploadable, "folder_to_upload_to": folder_id},
            toolkit="googledrive",
        )
        file_id = _find_value(data, ("id", "file_id"))
        if not file_id:
            raise GoogleWorkspaceError("Google Drive did not return an uploaded file id")
        return file_id

    async def delete_file(self, file_id: str) -> None:
        await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_DELETE_FILE",
            {"file_id": file_id},
            toolkit="googledrive",
        )

    async def create_folder(self, name: str, parent_id: str | None = None) -> str:
        arguments: dict[str, Any] = {"folder_name": name}
        if parent_id:
            arguments["parent_id"] = parent_id
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_CREATE_FOLDER",
            arguments,
            toolkit="googledrive",
        )
        folder_id = _find_value(data, ("id", "file_id", "folder_id"))
        if not folder_id:
            raise GoogleWorkspaceError("Google Drive did not return a folder id")
        return folder_id

    async def get_sheet_names(self, spreadsheet_id: str) -> list[str]:
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLESHEETS_GET_SHEET_NAMES",
            {"spreadsheet_id": spreadsheet_id},
            toolkit="googlesheets",
        )
        names = _find_string_list(data, ("sheet_names", "sheetNames", "sheets"))
        if not names:
            raise GoogleWorkspaceError("Google Sheets did not return worksheet names")
        return names

    async def download_file(self, file_id: str) -> bytes:
        data = await asyncio.to_thread(
            self._execute,
            "GOOGLEDRIVE_DOWNLOAD_FILE",
            {"file_id": file_id},
            toolkit="googledrive",
        )
        return await _read_download(data)

    async def create_report(
        self,
        job_profile: JobProfile,
        assessments: list[CandidateAssessment],
        profiles: list[CandidateProfile],
    ) -> ReportArtifacts:
        tables = build_report_tables(
            job_profile,
            assessments,
            profiles,
            public_app_url=self.settings.public_app_url,
        )
        title = f"Busca de Talentos - {job_profile.title} - {datetime.now(UTC):%Y%m%d-%H%M%S}"
        created = await asyncio.to_thread(
            self._execute,
            "GOOGLESHEETS_CREATE_GOOGLE_SHEET1",
            {"title": title},
            toolkit="googlesheets",
        )
        spreadsheet_id = _find_value(created, ("spreadsheetId", "spreadsheet_id", "id"))
        if not spreadsheet_id:
            raise GoogleWorkspaceError("Google Sheets did not return a spreadsheet id")
        sheet_url = _find_value(created, ("spreadsheetUrl", "spreadsheet_url", "url"))
        sheet_url = sheet_url or f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
        separator = formula_separator(_find_value(created, ("locale",)))

        default_sheet_id = _find_sheet_id(created)
        if default_sheet_id is None:
            spreadsheet_info = await asyncio.to_thread(
                self._execute,
                "GOOGLESHEETS_GET_SPREADSHEET_INFO",
                {"spreadsheet_id": spreadsheet_id},
                toolkit="googlesheets",
            )
            default_sheet_id = _find_sheet_id(spreadsheet_info)
        if default_sheet_id is None:
            raise GoogleWorkspaceError("Google Sheets did not return the default worksheet id")
        await asyncio.to_thread(
            self._execute,
            "GOOGLESHEETS_UPDATE_SHEET_PROPERTIES",
            {
                "spreadsheetId": spreadsheet_id,
                "updateSheetProperties": {
                    "properties": {"sheetId": default_sheet_id, "title": "Summary"},
                    "fields": "title",
                },
            },
            toolkit="googlesheets",
        )

        async def add_sheet(name: str) -> None:
            await asyncio.to_thread(
                self._execute,
                "GOOGLESHEETS_ADD_SHEET",
                {"spreadsheetId": spreadsheet_id, "properties": {"title": name}},
                toolkit="googlesheets",
            )
        await asyncio.gather(*(add_sheet(name) for name in ("Criteria", "Candidates")))

        table_map = {
            "Summary": tables.summary,
            "Criteria": tables.criteria,
            # Matches the XLSX, which already writes a labelled hyperlink. The
            # separator comes from the spreadsheet's own locale; when it is
            # unknown the bare URL is used, which stays clickable either way.
            "Candidates": [
                tables.candidates[0],
                *[
                    [row[0], candidate_cv_cell(cv_url, separator), *row[2:]]
                    for row, cv_url in zip(tables.candidates[1:], tables.candidate_cv_urls, strict=True)
                ],
            ],
        }

        async def update_sheet(name: str, values: list[list[Any]]) -> None:
            await asyncio.to_thread(
                self._execute,
                "GOOGLESHEETS_BATCH_UPDATE",
                {
                    "spreadsheet_id": spreadsheet_id,
                    "sheet_name": name,
                    "first_cell_location": "A1",
                    "valueInputOption": "USER_ENTERED",
                    "values": values,
                },
                toolkit="googlesheets",
            )
        await asyncio.gather(*(update_sheet(name, values) for name, values in table_map.items()))

        with tempfile.TemporaryDirectory(prefix="talent-search-") as temp_dir:
            xlsx_path = Path(temp_dir) / f"{title}.xlsx"
            await asyncio.to_thread(create_xlsx, xlsx_path, tables)
            uploadable = await asyncio.to_thread(
                self._stage_file,
                xlsx_path,
                slug="GOOGLEDRIVE_UPLOAD_FILE",
                toolkit="googledrive",
            )
            uploaded = await asyncio.to_thread(
                self._execute,
                "GOOGLEDRIVE_UPLOAD_FILE",
                {
                    "file_to_upload": uploadable,
                    "folder_to_upload_to": self.settings.results_drive_folder_id,
                },
                toolkit="googledrive",
            )
        file_id = _find_value(uploaded, ("id", "file_id"))
        if not file_id:
            raise GoogleWorkspaceError("Google Drive did not return an uploaded file id")
        excel_url = _find_value(uploaded, ("webViewLink", "web_view_link", "url", "file_url"))
        if not excel_url:
            excel_url = f"https://drive.google.com/open?id={file_id}"
        return ReportArtifacts(
            google_sheet_url=str(sheet_url),
            excel_url=str(excel_url),
            excel_file_id=str(file_id),
        )


def _find_file_records(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list) and all(isinstance(item, dict) for item in value):
        if any("id" in item or "file_id" in item for item in value):
            return value
    if isinstance(value, dict):
        for child in value.values():
            records = _find_file_records(child)
            if records:
                return records
    return []


def _find_value(value: Any, keys: tuple[str, ...]) -> str | None:
    if isinstance(value, dict):
        for key in keys:
            candidate = value.get(key)
            if isinstance(candidate, (str, int)) and str(candidate):
                return str(candidate)
        for child in value.values():
            found = _find_value(child, keys)
            if found:
                return found
    elif isinstance(value, list):
        for child in value:
            found = _find_value(child, keys)
            if found:
                return found
    return None


def _find_string_list(value: Any, keys: tuple[str, ...]) -> list[str]:
    if isinstance(value, dict):
        for key in keys:
            candidate = value.get(key)
            if isinstance(candidate, list) and all(isinstance(item, str) for item in candidate):
                return candidate
        for child in value.values():
            found = _find_string_list(child, keys)
            if found:
                return found
    elif isinstance(value, list):
        if all(isinstance(item, str) for item in value):
            return value
        for child in value:
            found = _find_string_list(child, keys)
            if found:
                return found
    return []


async def _read_download(value: Any) -> bytes:
    candidates: list[Any] = []

    def collect(item: Any) -> None:
        if isinstance(item, dict):
            for key, child in item.items():
                if key.lower() in {"content", "file", "path", "s3url", "url", "download_url"}:
                    candidates.append(child)
                collect(child)
        elif isinstance(item, list):
            for child in item:
                collect(child)

    collect(value)
    for candidate in candidates:
        if isinstance(candidate, bytes):
            return candidate
        if not isinstance(candidate, str) or not candidate:
            continue
        try:
            path = Path(candidate)
            if path.is_file():
                return await asyncio.to_thread(path.read_bytes)
        except OSError:
            pass
        if candidate.startswith(("http://", "https://")):
            async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
                response = await client.get(candidate)
                response.raise_for_status()
                return response.content
        try:
            decoded = base64.b64decode(candidate, validate=True)
            if decoded:
                return decoded
        except ValueError:
            continue
    raise GoogleWorkspaceError("Composio returned no readable file content")


def _find_sheet_id(value: Any) -> int | None:
    if isinstance(value, dict):
        for key in ("sheetId", "sheet_id"):
            candidate = value.get(key)
            if isinstance(candidate, int):
                return candidate
        for child in value.values():
            found = _find_sheet_id(child)
            if found is not None:
                return found
    elif isinstance(value, list):
        for child in value:
            found = _find_sheet_id(child)
            if found is not None:
                return found
    return None


# Google Sheets parses USER_ENTERED formulas with the spreadsheet's own locale,
# so the argument separator is not ours to choose. Locales are listed explicitly
# instead of guessed: an unknown one falls back to the bare URL, which is always
# clickable and can never render as #NAME?.
_SEMICOLON_LOCALES = frozenset(
    {"pt", "es", "de", "fr", "it", "nl", "ru", "tr", "pl", "id", "cs", "da", "fi", "nb", "sv", "uk", "ro", "hu"}
)
_COMMA_LOCALES = frozenset({"en", "ja", "ko", "zh", "he", "th", "ms"})


def formula_separator(locale: str | None) -> str | None:
    """Return the formula argument separator, or None when it cannot be trusted."""
    language = (locale or "").replace("-", "_").split("_")[0].casefold()
    if language in _SEMICOLON_LOCALES:
        return ";"
    if language in _COMMA_LOCALES:
        return ","
    return None


def candidate_cv_cell(url: str, separator: str | None) -> str:
    """Labelled hyperlink when the locale is known, plain URL otherwise."""
    if not url or separator is None:
        return url
    # The URL is built by us and percent-encoded, so it carries no quote to escape.
    return f'=HYPERLINK("{url}"{separator}"Currículo")'


def _safe_rows(rows: list[list[Any]]) -> list[list[Any]]:
    """Prevent resume/job text from becoming a spreadsheet formula."""
    return [[_safe_cell(cell) for cell in row] for row in rows]


def _safe_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
