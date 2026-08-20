from pathlib import Path

from openpyxl import load_workbook

from app.google_workspace import build_report_tables, create_xlsx
from app.schemas import CandidateAssessment, CandidateProfile, CriterionAssessment, JobCriterion, JobProfile


def _job() -> JobProfile:
    return JobProfile(
        title="Backend",
        criteria=[JobCriterion(id="python", description="Python", required=True, criterion_type="skill")],
    )


def _assessment(candidate_id: str, name: str) -> CandidateAssessment:
    return CandidateAssessment(
        candidate_id=candidate_id,
        candidate_name=name,
        criteria=[CriterionAssessment(criterion_id="python", status="supported", evidence=["Python"])],
        required_supported=1,
        required_total=1,
        criteria_coverage=1,
    )


def test_reports_have_three_tabs_alphabetical_candidates_and_no_formula_injection(tmp_path: Path) -> None:
    profiles = [
        CandidateProfile(
            candidate_id="2",
            full_name="Zeca",
            skills=["Python"],
            source_drive_file_id="2",
            source_drive_url="https://drive.test/2",
        ),
        CandidateProfile(
            candidate_id="1",
            full_name="Ana",
            skills=['=HYPERLINK("https://evil","click")'],
            source_drive_file_id="1",
            source_drive_url="https://drive.test/1",
        ),
    ]
    tables = build_report_tables(
        _job(),
        [_assessment("2", "Zeca"), _assessment("1", "Ana")],
        profiles,
    )
    destination = tmp_path / "report.xlsx"

    create_xlsx(destination, tables)

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == ["Summary", "Criteria", "Candidates"]
    candidates = workbook["Candidates"]
    assert [cell.value for cell in candidates[1]] == [
        "Candidato",
        "Currículo",
        "Formação",
        "Idiomas",
        "Experiência relevante (anos)",
        "Competências",
        "Critérios obrigatórios atendidos",
        "Total de critérios obrigatórios",
        "Cobertura dos critérios",
        "Evidências",
        "Pontos a confirmar",
        "Resumo profissional",
    ]
    criteria = workbook["Criteria"]
    assert [cell.value for cell in criteria[1]] == [
        "ID do critério",
        "Descrição",
        "Obrigatório",
        "Tipo de critério",
    ]
    assert [criteria.cell(2, column).value for column in range(1, 5)] == ["python", "Python", "Sim", "Competência"]
    assert candidates["A2"].value == "Ana"
    assert candidates["A3"].value == "Zeca"
    assert str(candidates["F2"].value).startswith("'=")
    assert candidates["B2"].value == "Abrir currículo"
    assert "python: Atendido" in str(candidates["J2"].value)
    assert candidates["B2"].hyperlink.target == "http://localhost:8000/api/v1/talent/candidates/1/cv"
    assert "drive.test" not in candidates["B2"].hyperlink.target
    assert tables.candidates[1][1] == "Abrir currículo"
    assert tables.candidate_cv_urls == [
        "http://localhost:8000/api/v1/talent/candidates/1/cv",
        "http://localhost:8000/api/v1/talent/candidates/2/cv",
    ]
    assert candidates.auto_filter.ref == candidates.dimensions
    assert candidates.sheet_view.showGridLines is False
